"""
Generate Azure Consumption Plan Excel for Universitas Terbuka (FY26-FY27)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ── Styling constants ──────────────────────────────────────────────────
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")  # Azure blue
SUB_HEADER_FILL = PatternFill(start_color="50A0E0", end_color="50A0E0", fill_type="solid")
TITLE_FONT = Font(name="Segoe UI", bold=True, size=14, color="0078D4")
SUBTITLE_FONT = Font(name="Segoe UI", bold=True, size=12, color="333333")
SECTION_FONT = Font(name="Segoe UI", bold=True, size=11, color="0078D4")
NORMAL_FONT = Font(name="Segoe UI", size=10)
BOLD_FONT = Font(name="Segoe UI", bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F5F9FF", end_color="F5F9FF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GROWTH_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
USD_FORMAT = '#,##0'
USD_FORMAT_DEC = '#,##0.00'
PCT_FORMAT = '0.0%'

def style_header_row(ws, row, max_col, fill=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data_cell(cell, is_currency=False, is_pct=False, bold=False):
    cell.font = BOLD_FONT if bold else NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right" if (is_currency or is_pct) else "left", vertical="center")
    if is_currency:
        cell.number_format = USD_FORMAT
    if is_pct:
        cell.number_format = PCT_FORMAT

def apply_total_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── DATA ───────────────────────────────────────────────────────────────
# Current FY26 ACR data (Jul 2025 – Apr 2026) from the provided report
MONTHS_FY26 = ["Jul-25", "Aug-25", "Sep-25", "Oct-25", "Nov-25", "Dec-25", "Jan-26", "Feb-26", "Mar-26", "Apr-26"]

SUBS = {
    "AZURE APP UT":             [0, 0, 0, 295, 377, 425, 420, 401, 574, 205],
    "AZURE E-LEARNING":         [1231, 1232, 4071, 45894, 44684, 17745, 14877, 46491, 90451, 29456],
    "AZURE IN OPEN":            [99988, 122782, 92467, 120836, 140956, 131736, 135660, 109095, 84438, 0],
    "AZURE SRS":                [0, 0, 0, 0, 0, 0, 4481, 8265, 9280, 3238],
    "Azure subscription 1":     [4051, 4112, 3943, 4166, 3601, 3675, 3761, 3467, 3763, 1058],
    "AZURE THE":                [0, 0, 80, 125, 122, 125, 125, 114, 125, 44],
    "DSI ELEARNING RND":        [0, 0, 157, 379, 557, 815, 1161, 756, 539, 155],
    "Microsoft Azure":          [2202, 2178, 2091, 2154, 2154, 2158, 2107, 1270, 1191, 419],
    "SUB-PLATFORM-NETWORK-IDC": [0, 0, 0, 0, 0, 0, 0, 251, 414, 145],
    "UNKNOWN":                  [105, 190, 243, 254, 262, 286, 182, 266, 192, 0],
}

MONTHS_PLAN = ["Apr-26", "May-26", "Jun-26", "Jul-26", "Aug-26", "Sep-26",
               "Oct-26", "Nov-26", "Dec-26", "Jan-27", "Feb-27", "Mar-27"]

GROWTH_RATE = 0.20  # 20% growth

# ── Reserved Instance: Ujian Online Platform ──────────────────────────
# Service estimates per region (monthly, USD, Pay-As-You-Go vs 1-Year RI)
RI_SERVICES = [
    # (Service, SKU/Description, PAYG/mo, RI 1yr/mo, Qty per region)
    ("Container Cluster (AKS)", "Standard_D4s_v5 (4 vCPU, 16 GB) x3 nodes", 620, 400, 3),
    ("Vue.js Frontend (App Service)", "P1v3 (2 vCPU, 8 GB)", 138, 90, 2),
    ("PHP Backend (Container)", "Standard_D2s_v5 (2 vCPU, 8 GB)", 310, 200, 2),
    ("PostgreSQL Flexible (Master)", "GP_Standard_D4s_v3 (4 vCPU, 16 GB)", 530, 345, 1),
    ("PostgreSQL Flexible (Slave/Replica)", "GP_Standard_D2s_v3 (2 vCPU, 8 GB)", 265, 172, 1),
    ("PostgreSQL Storage", "256 GB Premium SSD + Backup", 80, 80, 1),
    ("Azure CDN (Naskah Soal)", "Standard Microsoft CDN – 10 TB/mo egress", 850, 850, 1),
    ("Load Balancer", "Standard LB", 25, 25, 1),
    ("Azure Monitor & Log Analytics", "Basic tier", 60, 60, 1),
]

REGIONS = [
    ("Jakarta (Southeast Asia)", True),
    ("Bandung (Southeast Asia)", True),
    ("Semarang (Southeast Asia)", True),
    ("Lampung (Southeast Asia)", True),
]

# ── AI Course Generation Initiative ──────────────────────────────────
AI_SERVICES = [
    ("Azure OpenAI Service (GPT-4o)", "1M tokens/day avg", 4500),
    ("Azure OpenAI Service (Embeddings)", "ada-002, 2M tokens/day", 800),
    ("Azure AI Search", "S1 Standard (content indexing)", 750),
    ("Azure Cognitive Services (Speech)", "TTS for course narration", 400),
    ("Azure Blob Storage (Course Assets)", "Hot tier, 5 TB", 120),
    ("Azure Cosmos DB", "Autoscale 4000 RU/s (course metadata)", 580),
    ("Azure Container Apps", "AI processing workers (2 instances)", 450),
    ("Azure Redis Cache", "C2 Standard (caching)", 170),
    ("Azure App Service (AI Portal)", "P2v3", 280),
    ("Azure DevOps / CI-CD", "Pipeline agents", 150),
]

# ── Build workbook ────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# =====================================================================
# SHEET 1 – Executive Summary
# =====================================================================
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "0078D4"
set_col_widths(ws1, [3, 45, 22, 22, 22, 22, 22])

r = 2
ws1.cell(row=r, column=2, value="Azure Consumption Plan – Universitas Terbuka").font = TITLE_FONT
r += 1
ws1.cell(row=r, column=2, value="Consumption Plan April 2026 – March 2027 (12 Months)").font = SUBTITLE_FONT
r += 1
ws1.cell(row=r, column=2, value="Prepared: April 2026  |  Currency: USD").font = NORMAL_FONT
r += 2

# Summary table
headers = ["", "Metric", "FY26 Actual (10 mo)", "FY26 Annualized", "Plan Projected (20% Growth)"]
for ci, h in enumerate(headers, 1):
    ws1.cell(row=r, column=ci, value=h)
style_header_row(ws1, r, len(headers))
r += 1

fy26_actual = sum(sum(v) for v in SUBS.values())
# Annualize based on 9 full months (Jul-Mar), Apr is partial
fy26_9mo = sum(sum(v[:9]) for v in SUBS.values())
fy26_annualized = round(fy26_9mo / 9 * 12)

rows_data = [
    ("Current Azure Consumption", fy26_actual, fy26_annualized,
     round(fy26_annualized * (1 + GROWTH_RATE))),
]

# RI savings
total_ri_monthly = sum(svc[3] * svc[4] for svc in RI_SERVICES) * len(REGIONS)
total_payg_monthly = sum(svc[2] * svc[4] for svc in RI_SERVICES) * len(REGIONS)
ri_annual = total_ri_monthly * 12
payg_annual = total_payg_monthly * 12
ri_savings = payg_annual - ri_annual

ai_monthly = sum(s[2] for s in AI_SERVICES)
ai_annual = ai_monthly * 12

for label, actual, annual, projected in rows_data:
    ws1.cell(row=r, column=2, value=label); style_data_cell(ws1.cell(row=r, column=2))
    ws1.cell(row=r, column=3, value=actual); style_data_cell(ws1.cell(row=r, column=3), is_currency=True)
    ws1.cell(row=r, column=4, value=annual); style_data_cell(ws1.cell(row=r, column=4), is_currency=True)
    ws1.cell(row=r, column=5, value=projected); style_data_cell(ws1.cell(row=r, column=5), is_currency=True)
    r += 1

# New initiatives
ws1.cell(row=r, column=2, value="+ AI Course Generation (New)"); style_data_cell(ws1.cell(row=r, column=2))
ws1.cell(row=r, column=2).fill = GROWTH_FILL; ws1.cell(row=r, column=2).font = BOLD_FONT
ws1.cell(row=r, column=3, value="—"); style_data_cell(ws1.cell(row=r, column=3))
ws1.cell(row=r, column=4, value="—"); style_data_cell(ws1.cell(row=r, column=4))
ws1.cell(row=r, column=5, value=ai_annual); style_data_cell(ws1.cell(row=r, column=5), is_currency=True)
r += 1

ws1.cell(row=r, column=2, value="+ Ujian Online RI (New – 4 Regions)"); style_data_cell(ws1.cell(row=r, column=2))
ws1.cell(row=r, column=2).fill = GROWTH_FILL; ws1.cell(row=r, column=2).font = BOLD_FONT
ws1.cell(row=r, column=3, value="—"); style_data_cell(ws1.cell(row=r, column=3))
ws1.cell(row=r, column=4, value="—"); style_data_cell(ws1.cell(row=r, column=4))
ws1.cell(row=r, column=5, value=ri_annual); style_data_cell(ws1.cell(row=r, column=5), is_currency=True)
r += 1

ws1.cell(row=r, column=2, value="RI Savings vs Pay-As-You-Go"); style_data_cell(ws1.cell(row=r, column=2))
ws1.cell(row=r, column=2).fill = GREEN_FILL; ws1.cell(row=r, column=2).font = BOLD_FONT
ws1.cell(row=r, column=3, value="—"); style_data_cell(ws1.cell(row=r, column=3))
ws1.cell(row=r, column=4, value="—"); style_data_cell(ws1.cell(row=r, column=4))
ws1.cell(row=r, column=5, value=-ri_savings); style_data_cell(ws1.cell(row=r, column=5), is_currency=True)
r += 1

# Grand Total
total_plan = round(fy26_annualized * (1 + GROWTH_RATE)) + ai_annual + ri_annual
ws1.cell(row=r, column=2, value="GRAND TOTAL PROJECTED"); style_data_cell(ws1.cell(row=r, column=2), bold=True)
ws1.cell(row=r, column=5, value=total_plan); style_data_cell(ws1.cell(row=r, column=5), is_currency=True, bold=True)
apply_total_row(ws1, r, 5)
r += 2

# Key assumptions
ws1.cell(row=r, column=2, value="Key Assumptions & Notes").font = SECTION_FONT
r += 1
notes = [
    "1. FY26 data covers Jul 2025 – Apr 2026 (10 months). Annualized based on 9 full months (Apr partial).",
    "2. Plan period: April 2026 – March 2027 (12 months).",
    "3. Growth rate: 20% year-over-year applied per-month based on FY26 actuals.",
    "4. Apr-26 to Mar-27 projected from same calendar month in FY26 × (1 + 20%).",
    "5. Reserved Instance pricing based on 1-year commitment with estimated ~35% savings vs PAYG.",
    "6. Ujian Online Platform: 4 regions (JKT, BDG, SMG, Lampung), each with dedicated AKS cluster.",
    "7. AI Course Generation: new initiative covering GPT-4o, AI Search, Speech, and content storage.",
    "8. All prices in USD. Actual costs may vary based on usage patterns and Azure pricing changes.",
]
for note in notes:
    ws1.cell(row=r, column=2, value=note).font = NORMAL_FONT
    r += 1

# =====================================================================
# SHEET 2 – Current ACR by Subscription (FY26)
# =====================================================================
ws2 = wb.create_sheet("FY26 Current ACR")
ws2.sheet_properties.tabColor = "50A0E0"
set_col_widths(ws2, [3, 32] + [14]*10 + [16])

FY26_TITLE_ROW = 2
FY26_HEADER_ROW = 4
FY26_DATA_START = 5  # first subscription data row

r = FY26_TITLE_ROW
ws2.cell(row=r, column=2, value="FY26 ACR Details by Subscription (Monthly)").font = TITLE_FONT

r = FY26_HEADER_ROW
headers2 = ["", "Subscription"] + MONTHS_FY26 + ["Total"]
for ci, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=ci, value=h)
style_header_row(ws2, r, len(headers2))

r = FY26_DATA_START
sub_names_ordered = list(SUBS.keys())
for sub_name in sub_names_ordered:
    monthly_vals = SUBS[sub_name]
    ws2.cell(row=r, column=2, value=sub_name); style_data_cell(ws2.cell(row=r, column=2))
    for mi, val in enumerate(monthly_vals):
        cell = ws2.cell(row=r, column=3 + mi, value=val if val > 0 else 0)
        style_data_cell(cell, is_currency=True)
    # Total = SUM formula
    total_cell = ws2.cell(row=r, column=13)
    total_cell.value = f"=SUM(C{r}:L{r})"
    style_data_cell(total_cell, is_currency=True, bold=True)
    if r % 2 == 0:
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=r, column=c).fill = LIGHT_FILL
    r += 1

FY26_TOTAL_ROW = r
ws2.cell(row=r, column=2, value="TOTAL"); style_data_cell(ws2.cell(row=r, column=2), bold=True)
# Each month total = SUM formula
for mi in range(10):
    col_letter = get_column_letter(3 + mi)
    cell = ws2.cell(row=r, column=3 + mi)
    cell.value = f"=SUM({col_letter}{FY26_DATA_START}:{col_letter}{r - 1})"
    style_data_cell(cell, is_currency=True, bold=True)
# Grand total
ws2.cell(row=r, column=13).value = f"=SUM(M{FY26_DATA_START}:M{r - 1})"
style_data_cell(ws2.cell(row=r, column=13), is_currency=True, bold=True)
apply_total_row(ws2, r, len(headers2))

FY26_SHEET_NAME = "'FY26 Current ACR'"

# =====================================================================
# SHEET 3 – Projected FY27 Consumption by Subscription (FORMULA-DRIVEN)
# =====================================================================
ws3 = wb.create_sheet("Projected Apr26-Mar27")
ws3.sheet_properties.tabColor = "107C10"
set_col_widths(ws3, [3, 32] + [14]*12 + [16, 16, 14])

# Layout constants
FY27_TITLE_ROW = 2
FY27_GROWTH_ROW = 4
FY27_HEADER_ROW = 6
FY27_DATA_START = 7

r = FY27_TITLE_ROW
ws3.cell(row=r, column=2, value="Projected Consumption – April 2026 to March 2027").font = TITLE_FONT
r += 1
ws3.cell(row=r, column=2, value="Per-month projection: FY26 same month × (1 + Growth Rate)").font = SUBTITLE_FONT

# ── Growth Rate Parameter Cells (editable) ────────────────────────────
r = FY27_GROWTH_ROW
ws3.cell(row=r, column=2, value="Growth Rate:").font = BOLD_FONT
ws3.cell(row=r, column=2).alignment = Alignment(horizontal="right", vertical="center")
# C4 = the editable growth rate cell
GROWTH_CELL = f"$C${FY27_GROWTH_ROW}"  # absolute reference
ws3.cell(row=r, column=3, value=GROWTH_RATE)
ws3.cell(row=r, column=3).number_format = PCT_FORMAT
ws3.cell(row=r, column=3).font = Font(name="Segoe UI", bold=True, size=12, color="107C10")
ws3.cell(row=r, column=3).fill = GROWTH_FILL
ws3.cell(row=r, column=3).border = THIN_BORDER
ws3.cell(row=r, column=4, value="← Edit this cell to adjust all projections (current: 20%)")
ws3.cell(row=r, column=4).font = Font(name="Segoe UI", italic=True, size=10, color="888888")
ws3.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)

# ── Headers ───────────────────────────────────────────────────────────
r = FY27_HEADER_ROW
headers3 = ["", "Subscription"] + MONTHS_PLAN + ["Annual Total", "FY26 Total", "Growth %"]
for ci, h in enumerate(headers3, 1):
    ws3.cell(row=r, column=ci, value=h)
style_header_row(ws3, r, len(headers3))

# ── Data rows with formulas ──────────────────────────────────────────
# FY26 months: Jul(C) Aug(D) Sep(E) Oct(F) Nov(G) Dec(H) Jan(I) Feb(J) Mar(K) Apr(L)
# Plan cols:   Apr(C) May(D) Jun(E) Jul(F) Aug(G) Sep(H) Oct(I) Nov(J) Dec(K) Jan(L) Feb(M) Mar(N)
#              ─→L──  ─AVG── ─AVG── ─→C──  ─→D──  ─→E──  ─→F──  ─→G──  ─→H──  ─→I── ─→J──  ─→K──
# Apr-26 maps to FY26 Apr (col L)
# May-26 & Jun-26 have no prior-year month → use AVERAGE(FY26 all months)
# Jul-26 to Mar-27 map to the same calendar month in FY26

# Mapping: Plan column index (0-based) → FY26 column letter
# 0=Apr→L, 1=May→AVG, 2=Jun→AVG, 3=Jul→C, 4=Aug→D, 5=Sep→E,
# 6=Oct→F, 7=Nov→G, 8=Dec→H, 9=Jan→I, 10=Feb→J, 11=Mar→K
FY26_COL_MAP = ['L', None, None, 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

r = FY27_DATA_START
for si, sub_name in enumerate(sub_names_ordered):
    fy26_row = FY26_DATA_START + si  # corresponding row in FY26 sheet

    ws3.cell(row=r, column=2, value=sub_name); style_data_cell(ws3.cell(row=r, column=2))

    for mi in range(12):
        col = 3 + mi
        cell = ws3.cell(row=r, column=col)
        fy26_col = FY26_COL_MAP[mi]
        if fy26_col is None:
            # May-26 / Jun-26 → AVERAGE of all FY26 months
            cell.value = f"=AVERAGE({FY26_SHEET_NAME}!C{fy26_row}:L{fy26_row})*(1+{GROWTH_CELL})"
        else:
            # Direct month-to-month reference
            cell.value = f"={FY26_SHEET_NAME}!{fy26_col}{fy26_row}*(1+{GROWTH_CELL})"
        style_data_cell(cell, is_currency=True)

    # Annual Total (col O) = SUM of C:N
    cell_total = ws3.cell(row=r, column=15)
    cell_total.value = f"=SUM(C{r}:N{r})"
    style_data_cell(cell_total, is_currency=True, bold=True)

    # FY26 Total reference (col P) = reference FY26 Total column
    cell_fy26ref = ws3.cell(row=r, column=16)
    cell_fy26ref.value = f"={FY26_SHEET_NAME}!M{fy26_row}"
    style_data_cell(cell_fy26ref, is_currency=True)

    # Growth % (col Q) = (FY27 Annual / (FY26 Total * 12/10)) - 1
    # Or simply reference the growth rate cell
    cell_growth = ws3.cell(row=r, column=17)
    cell_growth.value = f"=IF(P{r}=0,\"\",O{r}/(P{r}*12/10)-1)"
    style_data_cell(cell_growth, is_pct=True)

    if r % 2 == 0:
        for c in range(1, len(headers3) + 1):
            ws3.cell(row=r, column=c).fill = LIGHT_FILL
    r += 1

FY27_TOTAL_ROW = r
FY27_DATA_END = r - 1

# ── Total row with SUM formulas ──────────────────────────────────────
ws3.cell(row=r, column=2, value="SUBTOTAL (Existing Subscriptions)"); style_data_cell(ws3.cell(row=r, column=2), bold=True)
# Each month total = SUM formula
for col in range(3, 15):  # C to N (12 months)
    col_letter = get_column_letter(col)
    cell = ws3.cell(row=r, column=col)
    cell.value = f"=SUM({col_letter}{FY27_DATA_START}:{col_letter}{FY27_DATA_END})"
    style_data_cell(cell, is_currency=True, bold=True)

# Annual Total
ws3.cell(row=r, column=15).value = f"=SUM(O{FY27_DATA_START}:O{FY27_DATA_END})"
style_data_cell(ws3.cell(row=r, column=15), is_currency=True, bold=True)
# FY26 Total
ws3.cell(row=r, column=16).value = f"=SUM(P{FY27_DATA_START}:P{FY27_DATA_END})"
style_data_cell(ws3.cell(row=r, column=16), is_currency=True, bold=True)
# Growth %
ws3.cell(row=r, column=17).value = f"=IF(P{r}=0,\"\",O{r}/(P{r}*12/10)-1)"
style_data_cell(ws3.cell(row=r, column=17), is_pct=True)
apply_total_row(ws3, r, len(headers3))

# ── Formula legend ───────────────────────────────────────────────────
r += 2
ws3.cell(row=r, column=2, value="Formula Reference:").font = SECTION_FONT
r += 1
formulas_legend = [
    "Apr-26:           = FY26 Apr × (1 + Growth Rate)",
    "May-26 & Jun-26:  = AVERAGE(FY26 Jul–Apr) × (1 + Growth Rate)  [no prior-year month available]",
    "Jul-26 to Mar-27: = [FY26 same month] × (1 + Growth Rate)  [e.g. Jul-26 = Jul-25 × (1+rate)]",
    "Annual Total:     = SUM(Apr-26 : Mar-27)",
    "Growth %:         = (Plan Annual Total / FY26 Annualized) - 1",
    f"Growth Rate cell: C{FY27_GROWTH_ROW} — change this value to recalculate all projections instantly.",
]
for note in formulas_legend:
    ws3.cell(row=r, column=2, value=note).font = NORMAL_FONT
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=14)
    r += 1

# =====================================================================
# SHEET 4 – AI Course Generation
# =====================================================================
ws4 = wb.create_sheet("AI Course Generation")
ws4.sheet_properties.tabColor = "E3008C"
set_col_widths(ws4, [3, 40, 35, 18, 18, 18])

r = 2
ws4.cell(row=r, column=2, value="Growth Initiative: AI Course Generation Platform").font = TITLE_FONT
r += 1
ws4.cell(row=r, column=2, value="New Azure services for AI-powered course content creation").font = SUBTITLE_FONT
r += 2

headers4 = ["", "Azure Service", "SKU / Description", "Monthly Cost (USD)", "Annual Cost (USD)", "Notes"]
for ci, h in enumerate(headers4, 1):
    ws4.cell(row=r, column=ci, value=h)
style_header_row(ws4, r, len(headers4))
r += 1

ai_notes = [
    "Course content generation",
    "Content vectorization",
    "Knowledge base indexing",
    "Audio narration for courses",
    "Video, PDF, SCORM assets",
    "Course catalog & metadata",
    "Batch processing workers",
    "Response caching layer",
    "Admin & instructor portal",
    "CI/CD automation",
]

for i, (svc, desc, monthly) in enumerate(AI_SERVICES):
    ws4.cell(row=r, column=2, value=svc); style_data_cell(ws4.cell(row=r, column=2))
    ws4.cell(row=r, column=3, value=desc); style_data_cell(ws4.cell(row=r, column=3))
    ws4.cell(row=r, column=4, value=monthly); style_data_cell(ws4.cell(row=r, column=4), is_currency=True)
    ws4.cell(row=r, column=5, value=monthly * 12); style_data_cell(ws4.cell(row=r, column=5), is_currency=True)
    ws4.cell(row=r, column=6, value=ai_notes[i]); style_data_cell(ws4.cell(row=r, column=6))
    if r % 2 == 0:
        for c in range(1, len(headers4) + 1):
            ws4.cell(row=r, column=c).fill = LIGHT_FILL
    r += 1

# Total
ws4.cell(row=r, column=2, value="TOTAL AI Course Generation"); style_data_cell(ws4.cell(row=r, column=2), bold=True)
ws4.cell(row=r, column=4, value=ai_monthly); style_data_cell(ws4.cell(row=r, column=4), is_currency=True, bold=True)
ws4.cell(row=r, column=5, value=ai_annual); style_data_cell(ws4.cell(row=r, column=5), is_currency=True, bold=True)
apply_total_row(ws4, r, len(headers4))

# =====================================================================
# SHEET 5 – Reserved Instance: Ujian Online Platform
# =====================================================================
ws5 = wb.create_sheet("RI - Ujian Online Platform")
ws5.sheet_properties.tabColor = "FF8C00"
set_col_widths(ws5, [3, 38, 38, 8, 16, 16, 16, 16])

r = 2
ws5.cell(row=r, column=2, value="Reserved Instance Plan: Ujian Online Platform").font = TITLE_FONT
r += 1
ws5.cell(row=r, column=2, value="1-Year Reserved Instance Commitment – Per Region Breakdown").font = SUBTITLE_FONT
r += 1
ws5.cell(row=r, column=2, value="Architecture: Vue.js Frontend | PHP Backend (Containerized) | PostgreSQL Master-Slave | CDN | AKS Clusters").font = Font(name="Segoe UI", italic=True, size=10, color="555555")
r += 2

# Per-service per-region table
headers5 = ["", "Service", "SKU / Configuration", "Qty", "PAYG/mo (USD)", "RI 1yr/mo (USD)", "Savings/mo (USD)", "Savings %"]
for ci, h in enumerate(headers5, 1):
    ws5.cell(row=r, column=ci, value=h)
style_header_row(ws5, r, len(headers5))
r += 1

total_payg = 0
total_ri = 0
for svc, desc, payg, ri, qty in RI_SERVICES:
    payg_total = payg * qty
    ri_total = ri * qty
    savings = payg_total - ri_total
    savings_pct = savings / payg_total if payg_total > 0 else 0

    ws5.cell(row=r, column=2, value=svc); style_data_cell(ws5.cell(row=r, column=2))
    ws5.cell(row=r, column=3, value=desc); style_data_cell(ws5.cell(row=r, column=3))
    ws5.cell(row=r, column=4, value=qty); style_data_cell(ws5.cell(row=r, column=4))
    ws5.cell(row=r, column=5, value=payg_total); style_data_cell(ws5.cell(row=r, column=5), is_currency=True)
    ws5.cell(row=r, column=6, value=ri_total); style_data_cell(ws5.cell(row=r, column=6), is_currency=True)
    ws5.cell(row=r, column=7, value=savings); style_data_cell(ws5.cell(row=r, column=7), is_currency=True)
    ws5.cell(row=r, column=7).fill = GREEN_FILL
    ws5.cell(row=r, column=8, value=savings_pct); style_data_cell(ws5.cell(row=r, column=8), is_pct=True)
    total_payg += payg_total
    total_ri += ri_total
    if r % 2 == 0:
        for c in range(1, len(headers5) + 1):
            ws5.cell(row=r, column=c).fill = LIGHT_FILL
    r += 1

# Per-region subtotal
ws5.cell(row=r, column=2, value="SUBTOTAL PER REGION (Monthly)"); style_data_cell(ws5.cell(row=r, column=2), bold=True)
ws5.cell(row=r, column=5, value=total_payg); style_data_cell(ws5.cell(row=r, column=5), is_currency=True, bold=True)
ws5.cell(row=r, column=6, value=total_ri); style_data_cell(ws5.cell(row=r, column=6), is_currency=True, bold=True)
ws5.cell(row=r, column=7, value=total_payg - total_ri); style_data_cell(ws5.cell(row=r, column=7), is_currency=True, bold=True)
ws5.cell(row=r, column=7).fill = GREEN_FILL
ws5.cell(row=r, column=8, value=(total_payg - total_ri) / total_payg); style_data_cell(ws5.cell(row=r, column=8), is_pct=True)
apply_total_row(ws5, r, len(headers5))
r += 2

# Multi-region summary
ws5.cell(row=r, column=2, value="Multi-Region Deployment Summary (4 Regions)").font = SECTION_FONT
r += 1
headers5b = ["", "Region", "Cluster Type", "PAYG/mo (USD)", "RI 1yr/mo (USD)", "RI Annual (USD)", "Savings Annual (USD)"]
for ci, h in enumerate(headers5b, 1):
    ws5.cell(row=r, column=ci, value=h)
style_header_row(ws5, r, len(headers5b))
r += 1

region_names = ["Jakarta (JKT)", "Bandung (BDG)", "Semarang (SMG)", "Lampung (LPG)"]
for rname in region_names:
    ws5.cell(row=r, column=2, value=rname); style_data_cell(ws5.cell(row=r, column=2))
    ws5.cell(row=r, column=3, value="Dedicated AKS Cluster"); style_data_cell(ws5.cell(row=r, column=3))
    ws5.cell(row=r, column=4, value=total_payg); style_data_cell(ws5.cell(row=r, column=4), is_currency=True)
    ws5.cell(row=r, column=5, value=total_ri); style_data_cell(ws5.cell(row=r, column=5), is_currency=True)
    ws5.cell(row=r, column=6, value=total_ri * 12); style_data_cell(ws5.cell(row=r, column=6), is_currency=True)
    ws5.cell(row=r, column=7, value=(total_payg - total_ri) * 12); style_data_cell(ws5.cell(row=r, column=7), is_currency=True)
    ws5.cell(row=r, column=7).fill = GREEN_FILL
    r += 1

# Grand total all regions
ws5.cell(row=r, column=2, value="GRAND TOTAL (4 Regions)"); style_data_cell(ws5.cell(row=r, column=2), bold=True)
ws5.cell(row=r, column=4, value=total_payg * 4); style_data_cell(ws5.cell(row=r, column=4), is_currency=True, bold=True)
ws5.cell(row=r, column=5, value=total_ri * 4); style_data_cell(ws5.cell(row=r, column=5), is_currency=True, bold=True)
ws5.cell(row=r, column=6, value=total_ri * 4 * 12); style_data_cell(ws5.cell(row=r, column=6), is_currency=True, bold=True)
ws5.cell(row=r, column=7, value=(total_payg - total_ri) * 4 * 12); style_data_cell(ws5.cell(row=r, column=7), is_currency=True, bold=True)
ws5.cell(row=r, column=7).fill = GREEN_FILL
apply_total_row(ws5, r, len(headers5b))

# =====================================================================
# SHEET 6 – Regional Deployment Architecture
# =====================================================================
ws6 = wb.create_sheet("Regional Architecture")
ws6.sheet_properties.tabColor = "8764B8"
set_col_widths(ws6, [3, 25, 30, 25, 25, 25, 25])

r = 2
ws6.cell(row=r, column=2, value="Regional Deployment Architecture – Ujian Online").font = TITLE_FONT
r += 2

headers6 = ["", "Component", "Jakarta (JKT)", "Bandung (BDG)", "Semarang (SMG)", "Lampung (LPG)"]
for ci, h in enumerate(headers6, 1):
    ws6.cell(row=r, column=ci, value=h)
style_header_row(ws6, r, len(headers6))
r += 1

arch_rows = [
    ("AKS Cluster", "3x D4s_v5 nodes", "3x D4s_v5 nodes", "3x D4s_v5 nodes", "3x D4s_v5 nodes"),
    ("Frontend (Vue.js)", "2x P1v3 App Svc", "2x P1v3 App Svc", "2x P1v3 App Svc", "2x P1v3 App Svc"),
    ("Backend (PHP)", "2x D2s_v5 Container", "2x D2s_v5 Container", "2x D2s_v5 Container", "2x D2s_v5 Container"),
    ("PostgreSQL Master", "GP_D4s_v3", "GP_D4s_v3", "GP_D4s_v3", "GP_D4s_v3"),
    ("PostgreSQL Slave", "GP_D2s_v3 (Read Replica)", "GP_D2s_v3 (Read Replica)", "GP_D2s_v3 (Read Replica)", "GP_D2s_v3 (Read Replica)"),
    ("Storage (DB)", "256 GB Premium SSD", "256 GB Premium SSD", "256 GB Premium SSD", "256 GB Premium SSD"),
    ("CDN (Naskah Soal)", "Shared – Azure CDN Standard", "Edge POP", "Edge POP", "Edge POP"),
    ("Load Balancer", "Standard LB", "Standard LB", "Standard LB", "Standard LB"),
    ("Monitoring", "Log Analytics + Metrics", "Log Analytics + Metrics", "Log Analytics + Metrics", "Log Analytics + Metrics"),
]

for comp, *regions in arch_rows:
    ws6.cell(row=r, column=2, value=comp); style_data_cell(ws6.cell(row=r, column=2), bold=True)
    for ci, region_val in enumerate(regions, 3):
        ws6.cell(row=r, column=ci, value=region_val); style_data_cell(ws6.cell(row=r, column=ci))
    if r % 2 == 0:
        for c in range(1, len(headers6) + 1):
            ws6.cell(row=r, column=c).fill = LIGHT_FILL
    r += 1

r += 1
ws6.cell(row=r, column=2, value="Network Architecture Notes:").font = SECTION_FONT
r += 1
net_notes = [
    "• Each region has its own dedicated AKS cluster for exam isolation and latency optimization.",
    "• PostgreSQL master-slave replication across regions for disaster recovery.",
    "• Azure CDN origin in Jakarta; edge POPs serve Bandung, Semarang, and Lampung.",
    "• Front Door or Traffic Manager routes traffic to nearest healthy cluster.",
    "• VNet peering between regional clusters for cross-region communication.",
    "• Private endpoints for PostgreSQL and storage to minimize data exposure.",
]
for note in net_notes:
    ws6.cell(row=r, column=2, value=note).font = NORMAL_FONT
    ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1

# =====================================================================
# SHEET 7 – FY27 Consolidated Budget
# =====================================================================
ws7 = wb.create_sheet("FY27 Consolidated Budget")
ws7.sheet_properties.tabColor = "107C10"
set_col_widths(ws7, [3, 45, 18, 18, 18, 18, 18])

r = 2
ws7.cell(row=r, column=2, value="FY27 Consolidated Azure Budget – Universitas Terbuka").font = TITLE_FONT
r += 1
ws7.cell(row=r, column=2, value="April 2026 – March 2027").font = SUBTITLE_FONT
r += 2

headers7 = ["", "Budget Category", "Monthly (Est.)", "Annual (20% Growth)", "% of Total"]
for ci, h in enumerate(headers7, 1):
    ws7.cell(row=r, column=ci, value=h)
style_header_row(ws7, r, len(headers7))
r += 1

# Reference FY27 Projected sheet total row for existing subscriptions
FY27_SHEET_REF = "'Projected Apr26-Mar27'"

budget_items = [
    ("Existing Subscriptions (Organic Growth)",
     f"={FY27_SHEET_REF}!O{FY27_TOTAL_ROW}/12",
     f"={FY27_SHEET_REF}!O{FY27_TOTAL_ROW}"),
    ("AI Course Generation (New Initiative)", ai_monthly, ai_annual),
    ("Ujian Online RI – Jakarta (JKT)", total_ri, total_ri * 12),
    ("Ujian Online RI – Bandung (BDG)", total_ri, total_ri * 12),
    ("Ujian Online RI – Semarang (SMG)", total_ri, total_ri * 12),
    ("Ujian Online RI – Lampung (LPG)", total_ri, total_ri * 12),
]

first_data_row = r
for label, m_est, a_est in budget_items:
    ws7.cell(row=r, column=2, value=label); style_data_cell(ws7.cell(row=r, column=2))
    ws7.cell(row=r, column=3, value=m_est); style_data_cell(ws7.cell(row=r, column=3), is_currency=True)
    ws7.cell(row=r, column=4, value=a_est); style_data_cell(ws7.cell(row=r, column=4), is_currency=True)
    if r % 2 == 0:
        for c in range(1, len(headers7) + 1):
            ws7.cell(row=r, column=c).fill = LIGHT_FILL
    r += 1

last_data_row = r - 1

# Fill % of total using formula
for ri in range(first_data_row, r):
    ws7.cell(row=ri, column=5).value = f"=IF(D{r}=0,\"\",D{ri}/D{r})"
    style_data_cell(ws7.cell(row=ri, column=5), is_pct=True)

# Grand total using SUM formulas
ws7.cell(row=r, column=2, value="GRAND TOTAL"); style_data_cell(ws7.cell(row=r, column=2), bold=True)
ws7.cell(row=r, column=3).value = f"=SUM(C{first_data_row}:C{last_data_row})"
style_data_cell(ws7.cell(row=r, column=3), is_currency=True, bold=True)
ws7.cell(row=r, column=4).value = f"=SUM(D{first_data_row}:D{last_data_row})"
style_data_cell(ws7.cell(row=r, column=4), is_currency=True, bold=True)
ws7.cell(row=r, column=5, value=1.0); style_data_cell(ws7.cell(row=r, column=5), is_pct=True)
apply_total_row(ws7, r, len(headers7))
r += 2

# YoY comparison
grand_total_row = r - 2  # the GRAND TOTAL FY27 row (before r+=2)
ws7.cell(row=r, column=2, value="Year-over-Year Comparison").font = SECTION_FONT
r += 1
ws7.cell(row=r, column=2, value="FY26 Actual (10 mo)"); style_data_cell(ws7.cell(row=r, column=2))
ws7.cell(row=r, column=4, value=fy26_actual); style_data_cell(ws7.cell(row=r, column=4), is_currency=True)
fy26_actual_row = r
r += 1
ws7.cell(row=r, column=2, value="FY26 Annualized (10→12 mo)"); style_data_cell(ws7.cell(row=r, column=2))
ws7.cell(row=r, column=4, value=fy26_annualized); style_data_cell(ws7.cell(row=r, column=4), is_currency=True)
fy26_annual_row = r
r += 1
ws7.cell(row=r, column=2, value="Plan Projected Total (Apr26-Mar27)"); style_data_cell(ws7.cell(row=r, column=2))
ws7.cell(row=r, column=4).value = f"=D{grand_total_row}"
style_data_cell(ws7.cell(row=r, column=4), is_currency=True)
fy27_proj_row = r
r += 1
ws7.cell(row=r, column=2, value="Net Increase vs FY26 Annualized"); style_data_cell(ws7.cell(row=r, column=2), bold=True)
ws7.cell(row=r, column=4).value = f"=D{fy27_proj_row}-D{fy26_annual_row}"
style_data_cell(ws7.cell(row=r, column=4), is_currency=True, bold=True)
ws7.cell(row=r, column=4).fill = ORANGE_FILL
ws7.cell(row=r, column=5).value = f"=IF(D{fy26_annual_row}=0,\"\",D{r}/D{fy26_annual_row})"
style_data_cell(ws7.cell(row=r, column=5), is_pct=True)

# ── Save ──────────────────────────────────────────────────────────────
output_path = r"c:\Users\teddysudewo\OneDrive - Microsoft\Desktop\AccelerateDevGHCopilot\consumption-plan-ut\Azure_Consumption_Plan_UT_Apr26-Mar27.xlsx"
wb.save(output_path)
print(f"✅ Excel saved to: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   FY26 Actual (10 mo): ${fy26_actual:,}")
print(f"   FY26 Annualized:     ${fy26_annualized:,}")
print(f"   Plan Total:            ${total_plan:,}")
print(f"   RI Annual Savings (4 regions): ${ri_savings:,}")
